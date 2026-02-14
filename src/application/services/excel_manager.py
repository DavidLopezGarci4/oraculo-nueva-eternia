
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from src.infrastructure.database_cloud import SessionCloud
from src.domain.models import CollectionItemModel, ProductModel
from src.core.logger import logger

class ExcelManager:
    """
    Gestor del Excel Bridge para sincronización de inventario.
    Mantiene el Excel local actualizado con los datos del Oráculo.
    
    IMPORTANTE: Usa openpyxl directamente para preservar:
    - Múltiples pestañas (sheets) con categorías
    - Formato de celdas (colores, fuentes, estilos)
    - Fila de título (Row 1) y cabeceras (Row 2)
    - Solo modifica la columna 'Adquirido' in-place
    """
    
    HEADER_ROW = 2      # Row 2 contiene las cabeceras
    DATA_START_ROW = 3   # Los datos empiezan en Row 3
    ADQ_COL = 1          # Columna A = Adquirido
    FIG_ID_COL = 17      # Columna Q = Figure ID
    NAME_COL = 2         # Columna B = Name
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path

    def _create_backup(self) -> str:
        """Crea un backup del Excel antes de modificarlo."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.dirname(self.excel_path)
        basename = os.path.splitext(os.path.basename(self.excel_path))[0]
        backup_path = os.path.join(backup_dir, f"{basename}_backup_{timestamp}.xlsx")
        shutil.copy2(self.excel_path, backup_path)
        logger.info(f"🛡️ Backup creado: {os.path.basename(backup_path)}")
        return backup_path

    def sync_acquisitions_from_db(self, user_id: int):
        """
        Lee la DB y actualiza el Excel local IN-PLACE usando openpyxl.
        Preserva todas las pestañas, formato y estructura.
        Match por Figure ID (columna Q) con fallback por nombre (columna B).
        Solo modifica la columna A (Adquirido): 'Sí' / 'No'.
        """
        if not os.path.exists(self.excel_path):
            logger.warning(f"⚠️ Excel no encontrado en {self.excel_path}. Operación cancelada.")
            return False

        try:
            logger.info(f"📊 Sincronizando Excel Bridge: {os.path.basename(self.excel_path)}...")
            
            # 1. Crear backup de seguridad
            self._create_backup()
            
            # 2. Obtener colección del usuario desde la DB
            with SessionCloud() as db:
                collection = db.query(CollectionItemModel).filter(
                    CollectionItemModel.owner_id == user_id
                ).all()
                # Mapa por Figure ID (clave primaria de match)
                id_map = {}
                name_map = {}
                for item in collection:
                    if item.product.figure_id:
                        id_map[str(item.product.figure_id)] = item.acquired
                    if item.product.name:
                        name_map[item.product.name.lower()] = item.acquired
            
            logger.info(f"📦 Colección cargada: {len(id_map)} items con Figure ID, {len(name_map)} por nombre.")
            
            # 3. Abrir el Excel con openpyxl (preserva formato y estructura)
            wb = load_workbook(self.excel_path)
            total_changes = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_changes = 0
                
                # Verificar que la estructura es la esperada
                header_adq = ws.cell(row=self.HEADER_ROW, column=self.ADQ_COL).value
                header_fig = ws.cell(row=self.HEADER_ROW, column=self.FIG_ID_COL).value
                
                if not header_adq or 'adquirido' not in str(header_adq).lower():
                    logger.warning(f"⚠️ Sheet '{sheet_name}': columna Adquirido no encontrada en A{self.HEADER_ROW}. Saltando.")
                    continue
                
                # Recorrer filas de datos (desde Row 3 hasta el final)
                for row_idx in range(self.DATA_START_ROW, ws.max_row + 1):
                    # Obtener Figure ID de la celda Q
                    fig_id_cell = ws.cell(row=row_idx, column=self.FIG_ID_COL)
                    name_cell = ws.cell(row=row_idx, column=self.NAME_COL)
                    adq_cell = ws.cell(row=row_idx, column=self.ADQ_COL)
                    
                    fig_id = str(fig_id_cell.value) if fig_id_cell.value is not None else None
                    name = str(name_cell.value).lower() if name_cell.value is not None else ""
                    
                    # Buscar en la DB: prioridad Figure ID, fallback nombre
                    acquired = None
                    if fig_id and fig_id in id_map:
                        acquired = id_map[fig_id]
                    elif name and name in name_map:
                        acquired = name_map[name]
                    
                    if acquired is not None:
                        new_val = 'Sí' if acquired else 'No'
                        old_val = str(adq_cell.value) if adq_cell.value else ''
                        
                        # Solo escribir si el valor cambió
                        if old_val != new_val:
                            adq_cell.value = new_val
                            sheet_changes += 1
                
                if sheet_changes > 0:
                    logger.info(f"  📝 {sheet_name}: {sheet_changes} cambios")
                total_changes += sheet_changes
            
            # 4. Guardar el workbook (preserva todo el formato)
            wb.save(self.excel_path)
            wb.close()
            
            logger.success(f"✅ Excel Bridge: {total_changes} cambios en {len(wb.sheetnames)} pestañas de {os.path.basename(self.excel_path)}.")
            return True

        except Exception as e:
            logger.error(f"❌ Error crítico en Excel Bridge: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

    def export_full_backup(self, user_id: int, output_path: str):
        """
        Genera un nuevo Excel completo con todos los datos del usuario como respaldo puro.
        """
        try:
            import pandas as pd
            with SessionCloud() as db:
                items = db.query(CollectionItemModel).filter(CollectionItemModel.owner_id == user_id).all()
                data = []
                for it in items:
                    data.append({
                        "ID": it.product.id,
                        "Producto": it.product.name,
                        "EAN": it.product.ean,
                        "Categoría": it.product.category,
                        "Adquirido": "SÍ" if it.acquired else "NO",
                        "Estado": it.condition,
                        "Precio Compra": it.purchase_price,
                        "Fecha": it.acquired_at
                    })
                
                df = pd.DataFrame(data)
                df.to_excel(output_path, index=False)
                logger.success(f"🛡️ Resguardo Humano generado: {output_path}")
                return True
        except Exception as e:
            logger.error(f"❌ Error al generar resguardo Excel: {e}")
            return False
