#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Scraper MOTU Origins -> Excel multi-hoja con hipervínculos a imágenes locales.
Ejecución directa en PyCharm o terminal.

- Rutas relativas al directorio del script:
    ./MOTU/              (Excel y datos)
    ./MOTU/images/       (descarga de imágenes)
- HTTP robusto: User-Agent realista (override por env MOTU_UA), reintentos, timeouts, pausas.
- Scraping tolerante a cambios menores.
- Mantiene el estado "Adquirido" desde un Excel previo si existe.
- Valida y colorea "Sí"/"No" en Excel, e inserta hipervínculos a imágenes locales.
- Nombres de hoja únicos (case-insensitive) con sufijos si hay colisiones.

Dependencias:
    pip install requests beautifulsoup4 pandas openpyxl xlsxwriter
"""

import os
import re
import time
import hashlib
import logging
import urllib.parse
from pathlib import Path
from typing import List, Tuple, Optional, Set

import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

# =========================
# Configuración y constantes
# =========================

CHECKLIST_URL = "https://www.actionfigure411.com/masters-of-the-universe/origins-checklist.php"
SITE_BASE = "https://www.actionfigure411.com/"
DESIRED_ORDER = [
    "Adquirido", "Name", "Wave", "Year", "Retail",
    "Imagen", "Image Path", "Image URL", "Detail Link"
]

REQUEST_TIMEOUT = 20.0     # segundos por petición
POLITE_SLEEP = 0.6         # pausa entre peticiones (cortesía)
MAX_RETRIES = 4            # reintentos para 429/5xx

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# =========================
# Utilidades de ruta
# =========================

def get_project_paths() -> Tuple[Path, Path, Path]:
    """
    Devuelve (project_root, excel_path, images_dir).
    project_root = carpeta de datos en data/MOTU
    """
    # Assuming src/collectors/personal_collection.py
    script_dir = Path(__file__).resolve().parent
    # Go up to src, then to project root
    project_base = script_dir.parent.parent
    
    project_root = project_base / "data" / "MOTU"
    images_dir = project_root / "images"
    excel_path = project_root / "lista_MOTU.xlsx"
    
    project_root.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)
    return project_root, excel_path, images_dir

# =========================
# Sesión HTTP robusta
# =========================

def build_session() -> requests.Session:
    """
    Sesión HTTP robusta:
    - User-Agent realista tipo Chrome por defecto.
    - Permite override vía variable de entorno MOTU_UA.
    - Cabeceras de navegador comunes.
    - Reintentos con backoff para 429/5xx y respeto de Retry-After.
    """
    default_ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    user_agent = os.getenv("MOTU_UA", default_ua)

    s = requests.Session()
    s.headers.update({
        "User-Agent": user_agent,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
        "Referer": SITE_BASE,
    })

    retries = Retry(
        total=MAX_RETRIES,
        connect=3,
        read=3,
        backoff_factor=0.8,  # 0.8, 1.6, 3.2, 6.4 ...
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD", "OPTIONS"],
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retries, pool_connections=20, pool_maxsize=20)
    s.mount("https://", adapter)
    s.mount("http://", adapter)

    return s

def polite_pause():
    """Pausa entre peticiones para no saturar el servidor."""
    time.sleep(POLITE_SLEEP)

def safe_get(session: requests.Session, url: str, **kwargs) -> Optional[requests.Response]:
    """GET con timeout y captura de errores; devuelve Response o None."""
    try:
        resp = session.get(url, timeout=REQUEST_TIMEOUT, **kwargs)
        if not resp.ok:
            logging.warning("GET no OK %s -> %s", url, resp.status_code)
            return None
        return resp
    except requests.RequestException as exc:
        logging.warning("Error de red en GET %s: %s", url, exc)
        return None

# =========================
# Scraping helpers
# =========================

def urljoin(base: str, href: str) -> str:
    return urllib.parse.urljoin(base, href)

def find_sections(soup: BeautifulSoup) -> List[Tuple[str, BeautifulSoup]]:
    """
    Encuentra secciones: (título, tabla) en la página índice.
    Selector robusto: busca h2, con o sin strong, y valida keywords.
    Soporta MÚLTIPLES tablas por sección (si están divididas).
    Retorna una lista plana de (titulo, tabla) donde el título se repite si hay varias tablas.
    """
    sections: List[Tuple[str, BeautifulSoup]] = []
    
    # Palabras clave para validar si un H2 es una sección válida de muñecos
    VALID_KEYWORDS = ["checklist", "origins", "masterverse", "wwe", "turtles", "crossover"]
    
    all_h2s = list(soup.find_all("h2"))
    
    def _is_valid_header(tag) -> Tuple[bool, str]:
        """Devuelve (True, Titulo) si es un header principal, o (False, None)."""
        txt = tag.get_text(strip=True)
        # 1. Strong tag
        s_tag = tag.find("strong")
        if s_tag:
            return True, s_tag.get_text(strip=True)
        # 2. Keywords
        if any(k in txt.lower() for k in VALID_KEYWORDS):
            return True, txt
        return False, None
    
    # Identificar indices de headers válidos
    valid_indices = []
    for i, h2 in enumerate(all_h2s):
        is_valid, title = _is_valid_header(h2)
        if is_valid:
            valid_indices.append((i, h2, title))
            
    # Procesar intervalos
    for idx_in_valid, (h2_idx, h2_node, title) in enumerate(valid_indices):
        # Determinar el nodo límite (el siguiente header válido)
        limit_node = None
        if idx_in_valid + 1 < len(valid_indices):
            limit_node = valid_indices[idx_in_valid + 1][1]
            
        # Buscar TODAS las tablas siguientes que estén antes del límite
        # find_all_next devuelve en orden de aparición en el documento
        candidate_tables = h2_node.find_all_next("table")
        
        tables_found_for_section = 0
        for tbl in candidate_tables:
            # Check si nos pasamos del límite
            if limit_node:
                # Comprobar posición: si tbl aparece DESPUÉS de limit_node en el source.
                # sourceline es aproximado, mejor usar lógica de 'sourceline' o comparar orden.
                # Una forma robusta en BS4 es ver si limit_node está en los 'previous_elements' de tbl? Lento.
                # Mejor: si limit_node aparece ANTES que tbl en el árbol parseado.
                # BS4 no da comparación directa > < de nodos fácilmente sin indexar.
                # Truco: usar .index(limit_node) es lento globalmente.
                pass
                
                # Check simplificado: string parsing index (hacky pero rápido)
                # O mejor: stop when we see a table that is inside or after the limit section.
                # Si 'tbl' está después de 'limit_node', paramos.
                # Asumimos que find_all_next itera en orden del documento.
                # Si encontramos limit_node en los parents de tbl? No.
                
                # Vamos a usar una heurística de posición basada en el orden de 'all_next'.
                # Iterar generator con cuidado.
                pass 
        
        # Estrategia de iteración manual (más segura que find_all_next masivo)
        curr = h2_node.next_element
        while curr:
            if curr == limit_node:
                break
            
            if curr.name == "table":
                sections.append((title, curr))
                tables_found_for_section += 1
                
            curr = curr.next_element
            
    return sections

def clean_headers(table: BeautifulSoup) -> List[str]:
    """Lee <th> y rellena vacíos como Unnamed_i."""
    headers = []
    for i, th in enumerate(table.find_all("th")):
        text = (th.get_text() or "").strip()
        headers.append(text if text else f"Unnamed_{i}")
    return headers

def extract_detail_link(name_cell: BeautifulSoup) -> Optional[str]:
    """
    Intenta obtener un enlace de detalle asociado a la fila.
    1) <a> dentro del primer <td>
    2) Enlace en un sibling cercano (<h3> o similar)
    """
    if not name_cell:
        return None
    a0 = name_cell.find("a", href=True)
    if a0:
        return a0["href"]
    sib = name_cell.next_sibling
    if sib and getattr(sib, "find", None):
        a1 = sib.find("a", href=True)
        if a1:
            return a1["href"]
    return None

def extract_image_url(detail_html: str, base: str) -> Optional[str]:
    """Localiza un <a data-fancybox ... href='...'> y devuelve URL absoluta."""
    dsoup = BeautifulSoup(detail_html, "html.parser")
    a_img = dsoup.select_one("a[data-fancybox][href]")
    if not a_img:
        return None
    href = a_img.get("href")
    if not href:
        return None
    return urljoin(base, href)

def download_image(session: requests.Session, url: str, dest_path: Path) -> bool:
    """Descarga binaria con streaming si no existe. Devuelve True si queda en disco."""
    try:
        if dest_path.exists():
            return True
        resp = session.get(url, stream=True, timeout=REQUEST_TIMEOUT)
        if not resp.ok:
            logging.warning("No se pudo descargar imagen %s -> %s", url, resp.status_code)
            return False
        with dest_path.open("wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except requests.RequestException as exc:
        logging.warning("Error descargando imagen %s: %s", url, exc)
        return False

def process_table(
    table: BeautifulSoup,
    session: requests.Session,
    images_dir: Path,
    fast_mode: bool = False,
    existing_keys: Set[str] = None
) -> pd.DataFrame:
    """
    Procesa una tabla HTML y retorna un DataFrame.
    fast_mode: Si es True, salta la descarga de detalles e imágenes.
    existing_keys: Conjunto de claves (Name|Year) que ya existen. Si está, se salta el detalle.
    """
    if existing_keys is None:
        existing_keys = set()
        
    headers = clean_headers(table)
    # AÑADIDO: 'Figure ID' para futuro uso robusto
    out_cols = ["Adquirido"] + headers + ["Detail Link", "Image URL", "Image Path", "Imagen", "Figure ID"]
    rows = []

    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue

        # Normaliza nº de celdas
        # Sanitizar smart quotes y espacios
        cells = []
        for td in tds:
            txt = (td.get_text() or "").strip()
            # Replace smart quotes
            txt = txt.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
            cells.append(txt)
            
        if len(cells) < len(headers):
            cells += [""] * (len(headers) - len(cells))
        else:
            cells = cells[:len(headers)]

        # --- Extracción de Figure ID (Análisis Robustez) ---
        figure_id = ""
        input_box = tr.find("input", {"type": "checkbox"})
        if input_box and input_box.has_attr("figureid"):
            figure_id = input_box["figureid"]

        # Construir clave actual para check incremental
        name_val = cells[0] if len(cells) > 0 else ""
        year_val = ""
        if len(cells) > 2:
            year_val = cells[2]
        
        current_key = f"{name_val}|{year_val}"
        
        is_existing = current_key in existing_keys
        should_skip_detail = fast_mode or is_existing

        row_data = ["No"] + cells

        name_cell = tds[0] if tds else None
        detail_link = extract_detail_link(name_cell)

        image_url = None
        image_path_str = None

        # Solo procesar detalles si NO debemos saltar
        if detail_link and not should_skip_detail:
            detail_url = urljoin(CHECKLIST_URL, detail_link)
            dresp = safe_get(session, detail_url)
            if dresp is not None:
                img_url = extract_image_url(dresp.text, SITE_BASE)
                if img_url:
                    image_url = img_url
                    image_name = Path(urllib.parse.urlparse(img_url).path).name
                    image_path = images_dir / image_name
                    if download_image(session, img_url, image_path):
                        image_path_str = str(image_path)
                polite_pause()
        elif is_existing and not fast_mode:
             # Si existe, NO descargamos
             pass

        row_data += [detail_link, image_url, image_path_str, "", figure_id]
        rows.append(row_data)

    df = pd.DataFrame(rows, columns=out_cols)

    # Elimina todas las columnas Unnamed_* si quedaron
    unnamed = [c for c in df.columns if c.startswith("Unnamed_")]
    if unnamed:
        df = df.drop(columns=unnamed)

    return df

# =========================
# Excel: lectura y fusión
# =========================

def sanitize_sheet_base(title: str) -> str:
    """
    Limpia el título para usar como base de nombre de hoja:
    - Sustituye caracteres inválidos.
    - Mantiene letras/números/espacios/guiones bajos.
    - Colapsa espacios a guiones bajos.
    - Recorta a 31 chars (luego se ajusta con sufijos si hay colisión).
    """
    # Reemplazos básicos
    t = title.replace("/", "-").replace("\\", "-").replace(":", "-")
    t = t.replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
    # Solo letras/números/espacios/guiones/guion_bajo
    t = re.sub(r"[^A-Za-z0-9 \-_]", "", t)
    # Colapsa espacios a "_"
    t = re.sub(r"\s+", "_", t.strip())
    if not t:
        t = "Sheet"
    return t[:31]  # base inicial (luego puede acortarse para sufijos)

def unique_sheet_name(title: str, used: Set[str]) -> str:
    """Genera un nombre de hoja único (case-insensitive) de <=31 caracteres."""
    base = sanitize_sheet_base(title)
    name = base
    norm = name.lower()
    if norm not in used:
        used.add(norm)
        return name

    # Añade sufijos incrementales
    counter = 2
    while True:
        suffix = f"_{counter}"
        maxlen = 31 - len(suffix)
        candidate = (base[:maxlen] if maxlen > 0 else base[:31]) + suffix
        norm_c = candidate.lower()
        if norm_c not in used:
            used.add(norm_c)
            return candidate
        counter += 1
        # Seguridad
        if counter > 99:
            h = hashlib.sha1(title.encode("utf-8")).hexdigest()[:4]
            base2 = (base[:31 - 5]) + "_" + h
            base = base2
            counter = 2

def read_existing_excel(file_path: Path) -> List[Tuple[str, pd.DataFrame]]:
    """Lee el Excel existente."""
    if not file_path.exists():
        logging.info("No existe Excel previo, se creará uno nuevo.")
        return []

    wb = load_workbook(file_path)
    sections_data_old: List[Tuple[str, pd.DataFrame]] = []

    for ws in wb.worksheets:
        title = ws.cell(row=1, column=1).value
        if not title:
            continue

        headers = []
        col = 1
        while True:
            val = ws.cell(row=2, column=col).value
            if val is None:
                break
            headers.append(val)
            col += 1

        data_rows = []
        r = 3
        max_row = ws.max_row
        while r <= max_row:
            row_values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            if all(x is None for x in row_values):
                break
            data_rows.append(row_values)
            r += 1

        df_section = pd.DataFrame(data_rows, columns=headers)
        sections_data_old.append((title, df_section))

    logging.info("Excel previo leído con éxito.")
    return sections_data_old

def make_key(df: pd.DataFrame) -> pd.DataFrame:
    """Crea una clave compuesta estable para el merge."""
    df = df.copy()
    candidates = [c for c in ["Name", "Wave", "Year"] if c in df.columns]
    if not candidates:
        df["__key__"] = df.index.astype(str)
        return df

    def _join_row(row):
        return "|".join([str(row[c]).strip() if pd.notna(row[c]) else "" for c in candidates])

    df["__key__"] = df.apply(_join_row, axis=1)
    return df

def combine_sections(
    sections_new: List[Tuple[str, pd.DataFrame]],
    sections_old: List[Tuple[str, pd.DataFrame]]
) -> List[Tuple[str, pd.DataFrame]]:
    """Combina datos manteniendo estado 'Adquirido'."""
    old_map = {t: df for t, df in sections_old}
    result: List[Tuple[str, pd.DataFrame]] = []

    def _dedupe_index(dfk: pd.DataFrame) -> pd.DataFrame:
        if not dfk.index.has_duplicates:
            return dfk
        counts = {}
        new_index = []
        for k in dfk.index:
            n = counts.get(k, 0)
            new_index.append(k if n == 0 else f"{k}#{n}")
            counts[k] = n + 1
        dfk = dfk.copy()
        dfk.index = new_index
        return dfk

    for title, df_new in sections_new:
        df_old = old_map.get(title, pd.DataFrame())

        # Alinea columnas
        new_cols = df_new.columns.tolist()
        old_cols = df_old.columns.tolist()
        missing_in_new = [c for c in old_cols if c not in new_cols]
        final_cols = new_cols + missing_in_new

        df_new = df_new.reindex(columns=final_cols, fill_value="")
        df_old = df_old.reindex(columns=final_cols, fill_value="")

        df_new_k = make_key(df_new).set_index("__key__", drop=False)
        df_old_k = make_key(df_old).set_index("__key__", drop=False)

        df_new_k = _dedupe_index(df_new_k)
        df_old_k = _dedupe_index(df_old_k)

        final_rows = []

        # Filas nuevas (Web)
        for k in df_new_k.index:
            if k in df_old_k.index:
                old_row = df_old_k.loc[k].copy()
                
                # INYECCIÓN DE FIGURE ID ROBUSTA
                # Si tenemos un ID nuevo (de la web), lo forzamos en la fila vieja
                # para ir migrando gradualmente a IDs sin perder datos viejos (imágenes, notas).
                if "Figure ID" in df_new_k.columns:
                     new_id = df_new_k.loc[k].get("Figure ID", "")
                     if new_id:
                         old_row["Figure ID"] = new_id
                
                # En modo incremental, df_new puede venir sin imágenes (saltado).
                # Por tanto, si existe en old, casi siempre preferimos old_row (con sus imágenes ya bajadas).
                # Solo actualizamos el status Adquirido si fuera logicamente posible, pero
                # aqui simplemente mantenemos el estado del excel previo.
                # Si el usuario quiere refrescar metadatos completos, debería borrar el Excel o filas especificas.
                final_rows.append(old_row)
            else:
                new_row = df_new_k.loc[k].copy()
                # Correct logic for NEW item: Adquirido = No
                new_row["Adquirido"] = "No"
                final_rows.append(new_row)

        # Filas viejas perdidas (que ya no están en la web?)
        for k in df_old_k.index:
            if k not in df_new_k.index:
                final_rows.append(df_old_k.loc[k].copy())

        df_final = pd.DataFrame(final_rows, columns=final_cols)

        existing_in_desired = [c for c in DESIRED_ORDER if c in df_final.columns]
        remaining_cols = [c for c in df_final.columns if c not in existing_in_desired]
        df_final = df_final.reindex(columns=existing_in_desired + remaining_cols)

        result.append((title, df_final))

    return result

# =========================
# Excel: escritura
# =========================

def write_excel_with_links(
    excel_path: Path,
    sections: List[Tuple[str, pd.DataFrame]]
) -> None:
    """Escribe Excel final."""
    with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
        wb = writer.book
        title_fmt = wb.add_format({"bg_color": "#ADD8E6", "bold": True})
        fmt_green = wb.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
        fmt_red = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})

        used_names: Set[str] = set()

        for title, df in sections:
            sheet_name = unique_sheet_name(title, used_names)
            ws = wb.add_worksheet(sheet_name)
            writer.sheets[sheet_name] = ws

            num_cols = max(1, df.shape[1])
            ws.merge_range(0, 0, 0, num_cols - 1, title, title_fmt)

            df.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=0, index=False)

            df_rows = df.shape[0]
            if df_rows == 0:
                continue

            if "Adquirido" in df.columns:
                col_idx = df.columns.get_loc("Adquirido")
                data_start = 2
                data_end = 1 + df_rows

                ws.data_validation(
                    data_start, col_idx, data_end, col_idx,
                    {"validate": "list", "source": ["Sí", "No"]}
                )

                col_letter = xlsxwriter.utility.xl_col_to_name(col_idx)
                data_range = f"{col_letter}{data_start+1}:{col_letter}{data_end+1}"

                ws.conditional_format(data_range, {
                    "type": "cell", "criteria": "==", "value": '"Sí"', "format": fmt_green
                })
                ws.conditional_format(data_range, {
                    "type": "cell", "criteria": "==", "value": '"No"', "format": fmt_red
                })

            if "Imagen" in df.columns and "Image Path" in df.columns:
                imagen_col = df.columns.get_loc("Imagen")
                path_col = df.columns.get_loc("Image Path")
                data_start = 2
                for i in range(df_rows):
                    img_path = df.iat[i, path_col]
                    if img_path:
                        p = Path(img_path)
                        if p.exists():
                            link = "file:///" + p.as_posix()
                            ws.write_url(data_start + i, imagen_col, link, string="Ver Imagen")

# =========================
# Reporting
# =========================

def generate_text_report(sections: List[Tuple[str, pd.DataFrame]], path: Path):
    """Genera un reporte de auditoría en texto."""
    total_items = 0
    with path.open("w", encoding="utf-8") as f:
        f.write("=== ACTIONFIGURE411 SCRAPING AUDIT REPORT ===\n")
        f.write(f"Generated at: {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        for title, df in sections:
            count = len(df)
            total_items += count
            f.write(f"SECTION: {title}\n")
            f.write(f"  Rows Found: {count}\n")
            if count > 0 and "Name" in df.columns:
                names = df["Name"].tolist()
                ids = df["Figure ID"].tolist() if "Figure ID" in df.columns else ["?"] * len(names)
                for n, fid in zip(names, ids):
                    f.write(f"    - [{fid}] {n}\n")
            f.write("-" * 40 + "\n")
            
        f.write(f"\nTOTAL ITEMS FOUND: {total_items}\n")
        
    logging.info(f"Audit report written to {path}")

# =========================
# Pipeline principal
# =========================

def main(audit_mode: bool = False) -> None:
    """
    Main entry point for scraping and updating the MOTU Catalog.
    Can be called directly from CLI or as a library (NexusService).
    """
    start_time = time.time()
    project_root, excel_path, images_dir = get_project_paths()
    
    session = build_session()
    
    # 1. Leer Excel PREVIO para incrementalidad
    existing_keys = set()
    if not audit_mode:
        logging.info("Leyendo Excel existente para modo incremental...")
        sections_old_preload = read_existing_excel(excel_path)
        for _, df_old in sections_old_preload:
            if "Name" in df_old.columns:
                 has_year = "Year" in df_old.columns
                 for _, row in df_old.iterrows():
                     n = str(row["Name"]).strip()
                     y = str(row["Year"]).strip() if has_year else ""
                     k = f"{n}|{y}"
                     existing_keys.add(k)
        logging.info("Se encontraron %d figuras existentes. Se omitirá su descarga.", len(existing_keys))

    logging.info("Descargando página índice...")
    resp = safe_get(session, CHECKLIST_URL)
    if resp is None:
        logging.error("No se pudo acceder a la página índice.")
        return
    soup = BeautifulSoup(resp.text, "html.parser")

    logging.info("Buscando secciones...")
    sections_html = find_sections(soup)
    logging.info("Secciones encontradas: %d", len(sections_html))

    # Procesar tablas
    sections_new: List[Tuple[str, pd.DataFrame]] = []
    for title, tbl in sections_html:
        logging.info("Procesando sección: %s (FastMode=%s, Incremental=%s)", title, audit_mode, not audit_mode)
        # Pasamos existing_keys
        df = process_table(tbl, session, images_dir, fast_mode=audit_mode, existing_keys=existing_keys)
        sections_new.append((title, df))
        if not audit_mode:
            polite_pause()

    if audit_mode:
        # Generate Report Only
        report_path = project_root.parent / "logs" / "scraping_report.txt"
        report_path.parent.mkdir(exist_ok=True)
        generate_text_report(sections_new, report_path)
        print(f"Report generated at: {report_path.absolute()}")
        return

    # Normal Flow: Merge and Write Excel
    # Re-leemos el Excel para merge robusto
    sections_old = read_existing_excel(excel_path)
    sections_final = combine_sections(sections_new, sections_old)
    write_excel_with_links(excel_path, sections_final)

    elapsed = time.time() - start_time
    logging.info("Extracción completada en %s", time.strftime("%H:%M:%S", time.gmtime(elapsed)))
    logging.info("Resultado guardado en: %s", excel_path)

def get_scraped_data() -> List[Tuple[str, pd.DataFrame]]:
    """Clean Architecture Entrypoint (Returns Data instead of writing to File)"""
    project_root, excel_path, images_dir = get_project_paths()
    session = build_session()
    
    resp = safe_get(session, CHECKLIST_URL)
    if resp is None: return []
        
    soup = BeautifulSoup(resp.text, "html.parser")
    sections_html = find_sections(soup)
    
    sections_data = []
    for title, tbl in sections_html:
        df = process_table(tbl, session, images_dir, fast_mode=False)
        sections_data.append((title, df))
        polite_pause()
        
    return sections_data

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="ActionFigures411 Scraper & Updater")
    parser.add_argument("--report", action="store_true", help="Run in audit mode (no downloads, generates report)")
    # Use parse_known_args to avoid crashing when called with other flags in multi-script environments
    args, unknown = parser.parse_known_args()
    
    main(audit_mode=args.report)