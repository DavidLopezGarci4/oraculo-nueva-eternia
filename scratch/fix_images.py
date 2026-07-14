import os
import requests
import pandas as pd
from PIL import Image
import io
import sqlite3

# 1. Directorios
images_dir = os.path.join('data', 'MOTU', 'images')
cache_dir = os.path.join('data', 'image_cache')
os.makedirs(images_dir, exist_ok=True)
os.makedirs(cache_dir, exist_ok=True)

# 2. Especificación de imágenes
targets = {
    13977: {
        'url': 'https://www.actionfigure411.com/masters-of-the-universe/images/man-at-arms-movie-13977.jpg',
        'slug': 'man-at-arms-movie-13977',
        'ext_excel': '.jpg'
    },
    13978: {
        'url': 'https://www.actionfigure411.com/masters-of-the-universe/images/tri-klops-movie-13978.jpg',
        'slug': 'tri-klops-movie-13978',
        'ext_excel': '.jpg'
    },
    14038: {
        'url': 'https://www.actionfigure411.com/masters-of-the-universe/images/skeletor-movie-14038.jpg',
        'slug': 'skeletor-movie-14038',
        'ext_excel': '.webp'
    },
    14039: {
        'url': 'https://www.actionfigure411.com/masters-of-the-universe/images/teela-camila-mendez-movie-14039.jpg',
        'slug': 'teela-camila-mendez-movie-14039',
        'ext_excel': '.webp'
    }
}

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

print("=== INICIANDO DESCARGA Y CONVERSIÓN DE IMÁGENES ===")

for fig_id, info in targets.items():
    url = info['url']
    slug = info['slug']
    ext_excel = info['ext_excel']
    
    print(f"\nDescargando ID #{fig_id} desde {url}...")
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code == 200:
            img_data = res.content
            # Abrir con Pillow para validar y convertir
            img = Image.open(io.BytesIO(img_data))
            
            # Guardar en data/MOTU/images con la extensión que pide el Excel
            if ext_excel == '.webp':
                dest_path = os.path.join(images_dir, f"{slug}.webp")
                img.save(dest_path, "WEBP", quality=85)
                print(f"  -> Guardada imagen local del catálogo: {dest_path}")
            else:
                dest_path = os.path.join(images_dir, f"{slug}.jpg")
                with open(dest_path, 'wb') as f:
                    f.write(img_data)
                print(f"  -> Guardada imagen local del catálogo: {dest_path}")
                
            # Guardar en data/image_cache como {id}.webp
            cache_path = os.path.join(cache_dir, f"{fig_id}.webp")
            img.save(cache_path, "WEBP", quality=85)
            print(f"  -> Guardada en caché del servidor local: {cache_path}")
            
        else:
            print(f"  ⚠️ Error de descarga (status {res.status_code}) para ID #{fig_id}")
    except Exception as e:
        print(f"  ❌ Fallo al procesar ID #{fig_id}: {e}")

# 3. Modificar Excel lista_MOTU.xlsx
excel_path = os.path.join('data', 'MOTU', 'lista_MOTU.xlsx')
if os.path.exists(excel_path):
    print("\n=== ACTUALIZANDO ARCHIVO EXCEL ===")
    try:
        # Cargar manteniendo formato/encabezado
        # El archivo excel tiene 1 fila de encabezado de título, y los datos empiezan en la fila 1 de cabeceras.
        df = pd.read_excel(excel_path, header=1)
        
        for fig_id, info in targets.items():
            slug = info['slug']
            ext_excel = info['ext_excel']
            url = info['url']
            
            # Localizar fila por Figure ID
            idx = df[df['Figure ID'] == fig_id].index
            if not idx.empty:
                # Modificar campos
                local_path_str = f"/home/runner/work/oraculo-nueva-eternia/oraculo-nueva-eternia/data/MOTU/images/{slug}{ext_excel}"
                df.loc[idx, 'Image Path'] = local_path_str
                df.loc[idx, 'Image URL'] = url
                df.loc[idx, 'Imagen'] = 'Ver Imagen'
                print(f"  -> Excel actualizado para ID #{fig_id}: {slug}{ext_excel}")
            else:
                print(f"  ⚠️ No se encontró la fila para ID #{fig_id} en el Excel")
                
        # Para restaurar la primera fila de título o guardar directamente con el formato
        # Re-escribir el excel en la hoja 'Origins' u hoja por defecto
        # Vamos a leer el archivo original completo con openpyxl para no destruir otras hojas si las hay,
        # o escribir de forma limpia respetando la fila de título 'Origins Action Figures Checklist'.
        # Para estar 100% seguros y mantener la fila 0 intacta, leeremos el archivo usando openpyxl.
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active # u obtener por nombre si es necesario
        
        # Buscar las columnas por sus nombres en la fila 2 (índice 2 en Excel)
        header_row = 2
        col_map = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=header_row, column=col).value
            if val:
                col_map[val] = col
                
        if 'Figure ID' in col_map and 'Image Path' in col_map and 'Image URL' in col_map:
            fig_id_col = col_map['Figure ID']
            img_path_col = col_map['Image Path']
            img_url_col = col_map['Image URL']
            
            # Buscar filas
            for row in range(header_row + 1, sheet.max_row + 1):
                val = sheet.cell(row=row, column=fig_id_col).value
                if val in targets:
                    info = targets[val]
                    slug = info['slug']
                    ext_excel = info['ext_excel']
                    url = info['url']
                    
                    local_path_str = f"/home/runner/work/oraculo-nueva-eternia/oraculo-nueva-eternia/data/MOTU/images/{slug}{ext_excel}"
                    sheet.cell(row=row, column=img_path_col).value = local_path_str
                    sheet.cell(row=row, column=img_url_col).value = url
                    print(f"  -> Excel (openpyxl) escrito para ID #{val}: {slug}{ext_excel}")
            wb.save(excel_path)
            print("  [OK] Excel guardado con exito preservando formato original.")
        else:
            print("  [ERROR] Columnas no encontradas en openpyxl")
    except Exception as e:
        print(f"  [ERROR] Error modificando el archivo Excel: {e}")
else:
    print(f"\n  [WARN] No existe {excel_path}")

# 4. Actualizar base de datos local oraculo.db
db_path = 'oraculo.db'
if os.path.exists(db_path):
    print("\n=== ACTUALIZANDO BASE DE DATOS LOCAL ===")
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Verificar si la tabla de productos tiene filas
        cursor.execute("SELECT COUNT(*) FROM products")
        count = cursor.fetchone()[0]
        
        if count > 0:
            for fig_id, info in targets.items():
                url = info['url']
                cursor.execute("UPDATE products SET image_url = ? WHERE figure_id = ?", (url, str(fig_id)))
                if cursor.rowcount > 0:
                    print(f"  -> DB local actualizada para ID #{fig_id}")
                else:
                    # Intentar buscar por id numérico directo por si acaso
                    cursor.execute("UPDATE products SET image_url = ? WHERE id = ?", (url, fig_id))
                    if cursor.rowcount > 0:
                        print(f"  -> DB local actualizada por id para ID #{fig_id}")
                    else:
                        print(f"  [WARN] ID #{fig_id} no se encontró en la tabla 'products' de la DB local")
            conn.commit()
            print("  [OK] Cambios en la DB local guardados con exito.")
        else:
            print("  [INFO] La tabla 'products' de la DB local esta vacia (0 registros). No requiere actualizacion inmediata.")
        conn.close()
    except Exception as e:
        print(f"  [ERROR] Error modificando base de datos local: {e}")

print("\n=== PROCESO COMPLETADO ===")
