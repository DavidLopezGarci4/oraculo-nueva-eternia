
import os
import sys
import pandas as pd
from pathlib import Path
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import logging
from datetime import datetime
from openpyxl import load_workbook
import shutil

# Add src to path
root_dir = Path(__file__).parent.parent
sys.path.append(str(root_dir))

from src.core.config import settings
from src.domain.models import ProductModel, CollectionItemModel, UserModel

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("ReverseSync")

def get_db_collection_status(session, username="David"):
    """
    Retrieves the collection status (Acquired: Yes/No) for all products for a specific user.
    Returns a dictionary mapping figure_id or (name|year) to 'Sí' or 'No'.
    """
    user = session.query(UserModel).filter_by(username=username).first()
    if not user:
        logger.warning(f"User {username} not found. Using empty collection.")
        return {}

    # Get all acquired product IDs
    acquired_product_ids = {
        item.product_id for item in session.query(CollectionItemModel).filter_by(
            owner_id=user.id, acquired=True
        ).all()
    }

    # Map products to status
    products = session.query(ProductModel).all()
    status_map = {}
    for p in products:
        status = "Sí" if p.id in acquired_product_ids else "No"
        
        # Use Figure ID as primary key if available
        if p.figure_id:
            status_map[str(p.figure_id).strip()] = status
        
        # Also use Name|Year as fallback/secondary key
        key_name_year = f"{p.name.strip()}|{p.release_year}"
        status_map[key_name_year] = status
    
    return status_map

def sync_excel_from_db(excel_path: str, status_map: dict):
    """
    Updates the 'Adquirido' column in the Excel file based on the status_map.
    Uses openpyxl to modify the existing file in-place, preserving formatting.
    """
    logger.info(f"🔄 Starting Reverse Sync for: {excel_path}")
    
    # Pre-flight backup
    backup_path = excel_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(excel_path, backup_path)
    logger.info(f"🛡️ Backup created: {backup_path}")

    try:
        wb = load_workbook(excel_path)
        changes_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"📄 Processing sheet: {sheet_name}")

            # Identify columns from header (row 2 as per our scrapers)
            headers = [str(ws.cell(row=2, column=c).value).strip() if ws.cell(row=2, column=c).value else f"Col{c}" for c in range(1, ws.max_column + 1)]
            
            try:
                idx_adquirido = headers.index("Adquirido") + 1
                idx_name = headers.index("Name") + 1
                idx_year = headers.index("Year") + 1
                idx_figure_id = headers.index("Figure ID") + 1 if "Figure ID" in headers else None
            except ValueError as e:
                logger.warning(f"⚠️ Missing columns in sheet {sheet_name}: {e}. Skipping.")
                continue

            # Process data rows starting from row 3
            for r in range(3, ws.max_row + 1):
                raw_name = ws.cell(row=r, column=idx_name).value
                raw_year = ws.cell(row=r, column=idx_year).value
                raw_fid = ws.cell(row=r, column=idx_figure_id).value if idx_figure_id else None
                
                if not raw_name:
                    continue

                # Clean values
                name = str(raw_name).strip()
                year = str(raw_year).strip()
                fid = str(raw_fid).strip() if raw_fid else None
                
                # Match logic: Figure ID (Priority) -> Name|Year
                new_status = None
                if fid and fid in status_map:
                    new_status = status_map[fid]
                else:
                    key = f"{name}|{year}"
                    if key in status_map:
                        new_status = status_map[key]

                if new_status:
                    current_status = ws.cell(row=r, column=idx_adquirido).value
                    if current_status != new_status:
                        ws.cell(row=r, column=idx_adquirido).value = new_status
                        changes_count += 1

        if changes_count > 0:
            wb.save(excel_path)
            logger.info(f"✅ Success: Updated {changes_count} status entries in {excel_path}")
        else:
            logger.info("ℹ️ No changes needed. Excel is already in sync with Database.")

    except Exception as e:
        logger.error(f"❌ Error during Excel update: {e}")
        raise
    finally:
        wb.close()

if __name__ == "__main__":
    from src.infrastructure.database_cloud import SessionCloud
    
    with SessionCloud() as db:
        # 1. Get status map from DB
        status_map = get_db_collection_status(db)
        
        # 2. Define Excel path
        project_root = Path(__file__).resolve().parent.parent
        excel_path = project_root / "data" / "MOTU" / "lista_MOTU.xlsx"
        
        if not excel_path.exists():
            logger.error(f"Target Excel not found: {excel_path}")
            sys.exit(1)
            
        # 3. Apply Sync
        sync_excel_from_db(str(excel_path), status_map)
